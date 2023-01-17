# Remington Zhu
# Jan 13, 2023
# Basic webscraping and excel spreadsheet creation.

import requests
from bs4 import BeautifulSoup
import csv
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
#importing libraries

def freeze(path, row_to_freeze):
    quote_list = []
    author_list = []
    #list of quotes and authors to be appended

    workbook = Workbook()
    #initialize Workbook

    sheet = workbook.active
    sheet.title = "ScrapedQuotes"
    sheet.freeze_panes = row_to_freeze
    sheet.column_dimensions['A'].width = 100
    sheet.column_dimensions['B'].width = 15
    headers = ["Quote", "Author"]
    sheet["A1"] = headers[0]
    sheet["B1"] = headers[1]
    # base layout of excel sheet, includes dimensions of columns and fill colours

    redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
    sheet['A1'].fill = redFill
    sheet['B1'].fill = redFill
    #custom red fill

    to_scrape = requests.get("http://quotes.toscrape.com") 
    soup = BeautifulSoup(to_scrape.text, "html.parser")
    #initialize the website to scrape data from

    quotes = soup.findAll("span", attrs={"class":"text"})
    authors = soup.findAll("small", attrs={"class":"author"})
    #finds all authors and quotes

    for quote in quotes:
        quote_list.append(str(quote.text))
    for author in authors:
        author_list.append(str(author.text))
    #print out each author and quote connected by a dash

    count = 2
    quoteauthorcount = 0
    for i in range(len(quote_list)):
        sheet.cell(row=count, column = 1).value = quote_list[quoteauthorcount]
        sheet.cell(row=count, column = 2).value = author_list[quoteauthorcount]
        count += 1
        quoteauthorcount += 1
    #for each row, print out value in quote_list and author_list

    workbook.save(path)

if __name__ == "__main__":
    freeze("ScrapedQuotes.xlsx", row_to_freeze="A2")
#"freezes" a cell so it can be visible later on




